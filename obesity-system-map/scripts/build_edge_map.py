"""Extract edge/node mapping from the geometry workbooks and verify it against
the actual SVG files, then emit JSON for the app to consume.

The key assumption under test: 'Links Layer Index' in the workbook equals the
document order of <path> elements inside the edges layer (excluding <defs>).
If that is wrong, every edge would be mislabelled, so it is checked exactly
against each path's `d` and `transform` rather than sampled.
"""

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(r"C:\Users\User\Desktop\NUS\HPB internship\Obesity\Obesity Systems Explorer v2")
EDGE_XLSX = ROOT / "obesity_system_map_complete_edge_geometry.xlsx"
NODE_XLSX = ROOT / "obesity_system_map_nodes_connections_with_box_boundaries.xlsx"
EDGES_SVG = ROOT / "obesity_map_edges_only.svg"
NODES_SVG = ROOT / "obesity_system_map_node_boxes_only.svg"
DATA_JSON = ROOT / "obesity_system_data.json"
OUT_DIR = ROOT / "obesity-system-map" / "src" / "data"

failures = []


def check(label, ok, detail=""):
    status = "PASS" if ok else "FAIL"
    print(f"[{status}] {label}" + (f" — {detail}" if detail else ""))
    if not ok:
        failures.append(label)
    return ok


def sheet_rows(wb, name):
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    return hdr, [r for r in it if any(c is not None for c in r)]


# ---------------------------------------------------------------- SVG parsing
def svg_paths_outside_defs(svg_text):
    """Return <path> tag strings in document order, excluding those in <defs>."""
    defs = re.search(r"<defs\b.*?</defs\s*>", svg_text, re.S)
    if defs:
        svg_text = svg_text[: defs.start()] + svg_text[defs.end():]
    return re.findall(r"<path\b[^>]*/?>", svg_text, re.S)


def attr(tag, name):
    m = re.search(rf'\b{name}="([^"]*)"', tag)
    return m.group(1) if m else None


def norm(s):
    return re.sub(r"\s+", " ", s).strip() if s else s


edges_svg = EDGES_SVG.read_text(encoding="utf-8")
edge_path_tags = svg_paths_outside_defs(edges_svg)
print(f"edges SVG: {len(edge_path_tags)} paths outside <defs>")

# ------------------------------------------------------------- workbook: edges
wb = openpyxl.load_workbook(EDGE_XLSX, read_only=True, data_only=True)
el_hdr, el_rows = sheet_rows(wb, "SVG Edge Elements")
eg_hdr, eg_rows = sheet_rows(wb, "Edge Geometry")

C = {name: el_hdr.index(name) for name in
     ["Element ID", "Links Layer Index", "Role", "Marker Type", "SVG d", "SVG Transform"]}
G = {name: eg_hdr.index(name) for name in
     ["Connection ID", "Source Node ID", "Target Node ID", "Influence",
      "All Line Element IDs", "Terminal Line Element", "Shared Line Elements",
      "Marker Fill Element", "Marker Outline Element", "Marker Type",
      "Geometry Mapping Note"]}

elements = sorted(el_rows, key=lambda r: int(r[C["Links Layer Index"]]))
check("element count == svg path count",
      len(elements) == len(edge_path_tags),
      f"{len(elements)} vs {len(edge_path_tags)}")

# ---- THE critical check: index ↔ document order, verified via d + transform
d_mismatch, t_mismatch = [], []
for i, (row, tag) in enumerate(zip(elements, edge_path_tags)):
    if int(row[C["Links Layer Index"]]) != i:
        d_mismatch.append((i, "index gap"))
        continue
    if norm(row[C["SVG d"]]) != norm(attr(tag, "d")):
        d_mismatch.append(i)
    if norm(row[C["SVG Transform"]]) != norm(attr(tag, "transform")):
        t_mismatch.append(i)

check("every path `d` matches workbook at same index", not d_mismatch,
      f"{len(d_mismatch)} mismatches {d_mismatch[:5]}")
check("every path `transform` matches workbook at same index", not t_mismatch,
      f"{len(t_mismatch)} mismatches {t_mismatch[:5]}")

# ---- role cross-check against rendered attributes
role_by_index = {int(r[C["Links Layer Index"]]): r[C["Role"]] for r in elements}
role_attr_mismatch = []
for i, tag in enumerate(edge_path_tags):
    role = role_by_index[i]
    fill, stroke = attr(tag, "fill"), attr(tag, "stroke")
    if role == "Line" and not (fill == "none" and stroke == "#231f20"):
        role_attr_mismatch.append((i, role, fill, stroke))
    if role == "Marker fill" and fill != "#231f20":
        role_attr_mismatch.append((i, role, fill, stroke))
check("role agrees with fill/stroke attributes", not role_attr_mismatch,
      f"{len(role_attr_mismatch)} odd {role_attr_mismatch[:3]}")

# ------------------------------------------------- element -> connection index
id_to_index = {r[C["Element ID"]]: int(r[C["Links Layer Index"]]) for r in elements}


def split_ids(cell):
    return [s.strip() for s in str(cell).split(",") if s and s.strip()] if cell else []


path_conns = defaultdict(list)   # path index -> [connection ids]
path_role = {}
path_marker = {}
conn_records = {}

unknown_refs = []
for r in eg_rows:
    cid = r[G["Connection ID"]]
    lines = split_ids(r[G["All Line Element IDs"]])
    terminal = r[G["Terminal Line Element"]]
    fill = r[G["Marker Fill Element"]]
    outline = r[G["Marker Outline Element"]]
    refs = [(e, "line") for e in lines]
    if fill:
        refs.append((fill, "markerFill"))
    if outline:
        refs.append((outline, "markerOutline"))
    idxs = []
    for eid, role in refs:
        if eid not in id_to_index:
            unknown_refs.append((cid, eid))
            continue
        i = id_to_index[eid]
        path_conns[i].append(cid)
        path_role[i] = role
        if role == "markerFill":
            path_marker[i] = r[G["Marker Type"]]
        idxs.append(i)
    note = r[G["Geometry Mapping Note"]]
    conn_records[cid] = {
        "sourceId": int(r[G["Source Node ID"]]),
        "targetId": int(r[G["Target Node ID"]]),
        "influence": str(r[G["Influence"]]).lower(),
        "markerType": r[G["Marker Type"]],
        "terminalIndex": id_to_index.get(terminal),
        "pathIndices": sorted(idxs),
        # "Direct endpoint match" is the clean case; other values mean the
        # attributed line does not start exactly on the source box.
        "mappingNote": note,
        "exactRoute": note == "Direct endpoint match",
    }

check("no unknown element references", not unknown_refs, str(unknown_refs[:5]))
check("Edge Geometry covers 296 connections", len(conn_records) == 296, str(len(conn_records)))
check("every connection has >=1 path", all(v["pathIndices"] for v in conn_records.values()))
check("every connection has a terminal index",
      all(v["terminalIndex"] is not None for v in conn_records.values()))

# ---------------------------------------------- cross-check against the JSON
data = json.loads(DATA_JSON.read_text(encoding="utf-8"))
json_conns = {c["id"]: c for c in data["connections"]}
json_nodes = {n["id"]: n for n in data["nodes"]}

check("connection ID sets identical (workbook vs JSON)",
      set(conn_records) == set(json_conns),
      f"only-wb={sorted(set(conn_records)-set(json_conns))[:5]} only-json={sorted(set(json_conns)-set(conn_records))[:5]}")

endpoint_mismatch = [
    cid for cid, v in conn_records.items()
    if cid in json_conns and (v["sourceId"] != json_conns[cid]["sourceId"]
                              or v["targetId"] != json_conns[cid]["targetId"])
]
check("source/target match JSON for every connection", not endpoint_mismatch,
      str(endpoint_mismatch[:5]))

influence_mismatch = [
    cid for cid, v in conn_records.items()
    if cid in json_conns and v["influence"] != json_conns[cid]["influence"]
]
check("influence matches JSON for every connection", not influence_mismatch,
      str(influence_mismatch[:5]))

marker_mismatch = [
    cid for cid, v in conn_records.items()
    if cid in json_conns
    and ((v["markerType"] == "Negative square") != (json_conns[cid]["sign"] == -1))
]
check("marker type agrees with JSON sign", not marker_mismatch, str(marker_mismatch[:5]))

# --------------------------------------------------------- orphan / shared
orphans = sorted(set(range(len(edge_path_tags))) - set(path_conns))
shared = {i: c for i, c in path_conns.items() if len(set(c)) > 1}
print(f"\norphan paths (in SVG, referenced by no connection): {len(orphans)} -> {orphans}")
for i in orphans:
    print(f"   index {i} ({'L%04d' % i}) role={role_by_index[i]!r} "
          f"dash={attr(edge_path_tags[i], 'stroke-dasharray')!r} "
          f"transform={attr(edge_path_tags[i], 'transform')!r}")
print(f"shared paths (one path serving multiple connections): {len(shared)}")
for i, c in list(shared.items())[:6]:
    print(f"   index {i} -> {sorted(set(c))}")

# ------------------------------------------------------------- workbook: nodes
wbn = openpyxl.load_workbook(NODE_XLSX, read_only=True, data_only=True)
nl_hdr, nl_rows = sheet_rows(wbn, "Node Box Layers")
N = {n: nl_hdr.index(n) for n in ["Layer ID", "Node ID", "Role"]}
layer_to_node = {r[N["Layer ID"]]: int(r[N["Node ID"]]) for r in nl_rows}

nodes_svg = NODES_SVG.read_text(encoding="utf-8")
svg_node_layers = re.findall(
    r'<path\b[^>]*\bid="([^"]+)"[^>]*\bdata-node-id="(\d+)"', nodes_svg)
check("node SVG layer count matches workbook",
      len(svg_node_layers) == len(layer_to_node),
      f"{len(svg_node_layers)} vs {len(layer_to_node)}")

svg_layer_map = {lid: int(nid) for lid, nid in svg_node_layers}
node_attr_mismatch = [
    lid for lid, nid in svg_layer_map.items()
    if layer_to_node.get(lid) != nid
]
check("every SVG data-node-id matches the workbook layer mapping",
      not node_attr_mismatch, str(node_attr_mismatch[:5]))
check("all 108 JSON nodes present in SVG",
      set(svg_layer_map.values()) == set(json_nodes),
      f"missing={sorted(set(json_nodes)-set(svg_layer_map.values()))[:5]}")

# ------------------------------------------------------------------- emit JSON
paths_out = []
for i in range(len(edge_path_tags)):
    conns = sorted(set(path_conns.get(i, [])))
    entry = {"role": path_role.get(i, {"Line": "line", "Marker fill": "markerFill",
                                       "Marker outline": "markerOutline"}[role_by_index[i]])}
    if conns:
        entry["connections"] = conns
    if i in path_marker:
        entry["markerType"] = path_marker[i]
    paths_out.append(entry)

out = {
    "_meta": {
        "source": "obesity_system_map_complete_edge_geometry.xlsx",
        "description": "Maps edge-layer <path> document order to connection IDs.",
        "pathCount": len(paths_out),
        "connectionCount": len(conn_records),
        "orphanPathIndices": orphans,
        # Connections whose attributed line does not begin on the source box.
        # Highlighting these lights up only the final leg of the drawn route.
        "approximateRouteConnections": sorted(
            cid for cid, v in conn_records.items() if not v["exactRoute"]
        ),
    },
    "paths": paths_out,
    "connections": {cid: conn_records[cid] for cid in sorted(conn_records)},
}
OUT_DIR.mkdir(parents=True, exist_ok=True)
(OUT_DIR / "edge_geometry.json").write_text(json.dumps(out, indent=1), encoding="utf-8")
print(f"\nwrote {OUT_DIR / 'edge_geometry.json'}")

# ------------------------------------------------------------ cluster legend
# Swatch colours are read from the node artwork rather than transcribed, so the
# legend cannot drift from the map. Several clusters contain one or two accent
# nodes (the highlighted hubs) plus a grey outline-ring fill, so the *modal*
# fill is what the printed legend shows.
LEGEND_ORDER = [
    "Media", "Social", "Psychological", "Economic", "Food",
    "Activity", "Infrastructure", "Developmental", "Biological", "Medical",
]

cluster_fills = defaultdict(Counter)
for match in re.finditer(
    r'<path[^>]*data-node-id="(\d+)"[^>]*fill="([^"]*)"', nodes_svg
):
    node_id, fill = int(match.group(1)), match.group(2)
    if fill != "none":
        cluster_fills[json_nodes[node_id]["mapCluster"]][fill] += 1

found = set(cluster_fills)
check("legend order covers every cluster", found == set(LEGEND_ORDER),
      f"missing={sorted(found - set(LEGEND_ORDER))} extra={sorted(set(LEGEND_ORDER) - found)}")

clusters_out = []
for name in LEGEND_ORDER:
    counts = cluster_fills[name]
    swatch, _ = counts.most_common(1)[0]
    members = [n["id"] for n in data["nodes"] if n["mapCluster"] == name]
    clusters_out.append({
        "name": name,
        "swatch": swatch,
        "nodeCount": len(members),
        "accentFills": sorted(f for f in counts if f != swatch),
    })

(OUT_DIR / "clusters.json").write_text(
    json.dumps({
        "_meta": {
            "source": "obesity_system_map_node_boxes_only.svg",
            "description": "Legend clusters with swatch colours read from the node artwork.",
            "order": "as printed on the original map legend",
        },
        "clusters": clusters_out,
    }, indent=1),
    encoding="utf-8",
)
print(f"wrote {OUT_DIR / 'clusters.json'}")

# --------------------------------------------------------- node box geometry
# Label text is wrapped to each node's real box width, so the boxes must come
# from the same source as the artwork rather than being estimated.
bg_hdr, bg_rows = sheet_rows(wbn, "Node Box Geometry")
B = {n: bg_hdr.index(n) for n in
     ["Node ID", "Original Map X", "Original Map Y", "Boundary Path Left X",
      "Boundary Path Top Y", "Boundary Path Width", "Boundary Path Height"]}

boxes = {}
for r in bg_rows:
    node_id = int(r[B["Node ID"]])
    boxes[node_id] = {
        "id": node_id,
        "x": round(float(r[B["Boundary Path Left X"]]), 3),
        "y": round(float(r[B["Boundary Path Top Y"]]), 3),
        "w": round(float(r[B["Boundary Path Width"]]), 3),
        "h": round(float(r[B["Boundary Path Height"]]), 3),
    }

check("a box for every node", set(boxes) == set(json_nodes),
      f"missing={sorted(set(json_nodes) - set(boxes))[:5]}")

# Box centre must agree with the node's map coordinate, else labels sit off-box.
centre_drift = []
for node_id, box in boxes.items():
    node = json_nodes[node_id]
    dx = abs(box["x"] + box["w"] / 2 - node["x"])
    dy = abs(box["y"] + box["h"] / 2 - node["y"])
    if dx > 1.5 or dy > 1.5:
        centre_drift.append((node_id, round(dx, 2), round(dy, 2)))
check("box centres match node x/y", not centre_drift, str(centre_drift[:5]))

(OUT_DIR / "node_boxes.json").write_text(
    json.dumps({
        "_meta": {
            "source": "obesity_system_map_nodes_connections_with_box_boundaries.xlsx",
            "description": "Node box bounds in map coordinates, for label layout.",
        },
        "boxes": [boxes[i] for i in sorted(boxes)],
    }, indent=1),
    encoding="utf-8",
)
print(f"wrote {OUT_DIR / 'node_boxes.json'}")
widths = [b["w"] for b in boxes.values()]
heights = [b["h"] for b in boxes.values()]
print(f"   box width  {min(widths):.1f}–{max(widths):.1f}")
print(f"   box height {min(heights):.1f}–{max(heights):.1f}")
for c in clusters_out:
    print(f"   {c['name']:16} {c['swatch']}  n={c['nodeCount']}")

print("\nrole distribution:", Counter(p["role"] for p in paths_out))
print("paths with connections:", sum(1 for p in paths_out if "connections" in p))
print(f"\n{'ALL CHECKS PASSED' if not failures else 'FAILURES: ' + ', '.join(failures)}")
sys.exit(1 if failures else 0)
