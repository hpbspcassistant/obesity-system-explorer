"""
Turns the programme-node remapping spreadsheet into src/data/intervention/programmes.json.

Run it again whenever the spreadsheet changes:

    python tools/import_inventory.py ../programme-node-remapping.xlsx [reasons.json]

An optional second argument is a JSON file keyed by programme name, whose values
are {nodeId: reason} objects.  When present, its reasons override column K.

Columns used (0-indexed):

  0  Programme name
  6  FY26/27 workplan check status  -> derives `status`
  7  Proposed nodes                 -> "71 Physical Activity\n49 ..."
  9  Gate (structured)              -> parsed into machine gate
 10  Reason (tool display)          -> "71 Physical Activity — reason\n..."

The script refuses to write anything it cannot fully account for: an unknown
gate, a node id in the reason column that was not in the proposed-nodes column,
or a missing reason for a proposed node all abort the run.
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "src" / "data" / "intervention"

# ----------------------------------------------------------------- gates

def parse_gate(text):
    """Turn the structured gate string into a machine gate."""
    text = (text or "").strip()
    if not text or text == "(sunsetted)":
        return "everyone"

    if re.match(r"^everyone", text, re.I):
        return "everyone"

    # Handle OR-joined clauses: "role=X OR age_band ∈ {a,b} OR ..."
    if " OR " in text:
        clauses = [_parse_single_clause(part.strip()) for part in text.split(" OR ")]
        return clauses

    return [_parse_single_clause(text)]


def _parse_single_clause(text):
    """Parse 'role=working AND work_type=manual-industrial' into a dict."""
    # Strip trailing bracket comments from the whole clause first
    text = re.sub(r"\s*\[.*?\]\s*$", "", text).strip()

    clause = {}
    parts = re.split(r"\s+AND\s+", text)
    for part in parts:
        part = part.strip()
        # Strip trailing parenthetical comments
        part = re.sub(r"\s*\(.*\)\s*$", "", part).strip()
        # Strip trailing bracket comments
        part = re.sub(r"\s*\[.*?\]\s*$", "", part).strip()
        if not part:
            continue

        # field ∈ {val1, val2, ...}
        m = re.match(r"^(\w+)\s*[∈]\s*\{([^}]+)\}", part)
        if m:
            field = m.group(1)
            values = [v.strip() for v in m.group(2).split(",")]
            clause[field] = _coerce_values(values) if len(values) > 1 else _coerce_single(values[0])
            continue

        # condition=X -> conditions: [X] (must precede generic field=value)
        m = re.match(r"^condition\s*=\s*([a-z0-9-]+)", part)
        if m:
            clause["conditions"] = [m.group(1).strip()]
            continue

        # field=value
        m = re.match(r"^(\w+)\s*=\s*([a-z0-9-]+)", part)
        if m:
            field = m.group(1)
            value = m.group(2).strip()
            clause[field] = _coerce_single(value)
            continue

        raise SystemExit(f"Cannot parse gate clause part: {part!r}")

    return clause


def _coerce_single(value):
    if value == "true":
        return True
    if value == "false":
        return False
    return value


def _coerce_values(values):
    return [_coerce_single(v) for v in values]


# ----------------------------------------------------------------- nodes

def parse_nodes(text):
    """Parse '71 Physical Activity\\n49 Level of ...' into a list of node ids."""
    text = (text or "").strip()
    if not text or text.startswith("("):
        return []
    ids = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(\d+)\s+", line)
        if m:
            ids.append(int(m.group(1)))
    return ids


def parse_reasons(text):
    """Parse '71 Physical Activity — reason\\n49 ...' into {node_id: reason}."""
    text = (text or "").strip()
    if not text or text.startswith("("):
        return {}
    reasons = {}
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(\d+)\s+[^—→]+[—→]\s*(.+)$", line)
        if m:
            reasons[int(m.group(1))] = m.group(2).strip()
    return reasons


# ----------------------------------------------------------------- status

def derive_status(text):
    """Map the workplan check column to a programme status."""
    text = (text or "").strip().lower()
    if "sunsetted" in text or "sunset" in text:
        return "ended"
    if "accurate" in text:
        return "current"
    # Everything else (verify, check, transition, not in workplan, etc.)
    return "verify"


# ----------------------------------------------------------------- slug

def slug(name):
    """A stable id from the programme name."""
    plain = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    plain = re.sub(r"\([^)]*\)", " ", plain)
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", plain.lower())).strip("-")


# ----------------------------------------------------------------- main

def main():
    source = Path(
        sys.argv[1]
        if len(sys.argv) > 1
        else HERE.parent.parent / "programme-node-remapping.xlsx"
    )

    reasons_override = {}
    if len(sys.argv) > 2:
        reasons_path = Path(sys.argv[2])
        raw = json.loads(reasons_path.read_text("utf-8"))
        for prog_name, entries in raw.items():
            reasons_override[prog_name] = {int(k): v for k, v in entries.items()}
        print(f"Loaded reason overrides for {len(reasons_override)} programmes")

    wb = openpyxl.load_workbook(source, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header

    out, seen = [], set()
    for row in rows:
        name = (row[0] or "").strip()
        if not name:
            continue

        status_text = str(row[6] or "")
        status = derive_status(status_text)

        # Skip sunsetted programmes entirely
        nodes_text = str(row[7] or "").strip()
        if nodes_text.startswith("(sunsetted)") or status == "ended":
            continue

        node_ids = parse_nodes(row[7])
        reasons = reasons_override.get(name) or parse_reasons(row[10])

        # Programmes with (none) reach no nodes — still include them
        gate_text = str(row[9] or "").strip()
        if gate_text.startswith("(sunsetted)"):
            continue

        identifier = slug(name)
        if identifier in seen:
            raise SystemExit(f"Duplicate id {identifier!r} from {name!r}")
        seen.add(identifier)

        gate = parse_gate(gate_text)

        # Build the nodes array with reasons
        nodes_with_reasons = []
        for nid in node_ids:
            reason = reasons.get(nid)
            if reason is None:
                raise SystemExit(
                    f"{name}: node {nid} is in proposed-nodes but has no reason"
                )
            nodes_with_reasons.append({"id": nid, "reason": reason})

        # Check for reasons referencing nodes not in the proposed list
        extra_reason_ids = set(reasons.keys()) - set(node_ids)
        if extra_reason_ids:
            raise SystemExit(
                f"{name}: reason references nodes {sorted(extra_reason_ids)} "
                f"not in proposed-nodes {node_ids}"
            )

        programme = {
            "id": identifier,
            "name": name,
            "source": str(row[3] or "").strip(),
            "status": status,
            "gate": gate,
            "gateSource": str(row[1] or "").strip(),
            "nodes": nodes_with_reasons,
        }
        out.append(programme)

    (DATA / "programmes.json").write_text(
        json.dumps(out, indent=2, ensure_ascii=False) + "\n", "utf-8"
    )

    reached = set()
    for programme in out:
        if programme["status"] == "ended":
            continue
        for node in programme["nodes"]:
            reached.add(node["id"])
    print(f"{len(out)} programmes -> {DATA / 'programmes.json'}")
    print(f"{len(reached)} nodes reached with gates ignored")


if __name__ == "__main__":
    main()
